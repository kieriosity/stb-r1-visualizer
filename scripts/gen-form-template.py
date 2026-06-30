"""Extract the STB R-1 xlsx form template into a compact JSON grid the web
viewer renders as a near-pixel facsimile. Static labels, instruction text,
column widths, borders, fonts and alignment come straight from the official
template; the viewer overlays each submission's numeric values into the
empty bordered value cells.

Run:
    python scripts/gen-form-template.py
Writes r1_visualizer/src/formTemplate.json
"""
import json, re, os, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / 'r1_visualizer' / 'src' / 'formTemplate.json'

# The STB renumbered Schedule 200 (and revised 210, added 210A) between these
# revisions, so the facsimile is generated per form version and the viewer renders
# whichever matches the filing's envelope.form_metadata.form_version. Identical
# pages are emitted once and tagged with both versions to keep the bundle small.
# Chronological order; the LAST entry is the canonical page order for the nav.
FORMS = [
    ('2015-08-31', ROOT / 'forms' / 'R1-08-31-2015.xlsx'),
    ('2026-07-31', ROOT / 'forms' / 'R1-7-31-2026.xlsx'),
]

# Sheet name -> data schedule_id (None = front matter / instructions, shown
# static). Keeps form order via the workbook's own sheet order.
def sheet_to_schedule(name):
    ids = sheet_schedules(name)
    return ids[0] if ids else None


def sheet_schedules(name):
    """All schedule ids named in a sheet, in order. The legacy form stacks two
    schedules on one sheet ('340 350', '342 351', '501 502'); each gets its own
    page (split at its banner). Instruction/front-matter sheets return []."""
    n = name.strip()
    if re.search(r'inst', n, re.I):
        return []
    ids = []
    for m in re.finditer(r'(PTC)?\s*\b([0-9]{3}[a-zA-Z]?)\b', n):
        sid = f'PTC_{m.group(2).upper()}' if m.group(1) else m.group(2).upper()
        if sid not in ids:
            ids.append(sid)
    return ids


def _split_combined(grid, ids):
    """Split a combined sheet's grid into one page per schedule, sliced at each
    schedule's '<id>.' banner row. Ids without a banner in the grid are dropped."""
    rows = grid['rows']
    found = []
    for sid in ids:
        pat = re.compile(rf'^\s*{re.escape(sid)}\.')
        idx = next((i for i, row in enumerate(rows)
                    if any(isinstance(c.get('t'), str) and pat.match(c['t']) for c in row['cells'])),
                   None)
        if idx is not None:
            found.append((idx, sid))
    found.sort()
    if len(found) <= 1:                    # nothing to split on; one page as-is
        out = dict(grid)
        out['schedule'] = found[0][1] if found else (ids[0] if ids else None)
        return [out]
    pages = []
    for k, (idx, sid) in enumerate(found):
        start = idx if k > 0 else 0         # first page keeps the leading page header
        end = found[k + 1][0] if k + 1 < len(found) else len(rows)
        sub = dict(grid)
        sub['rows'] = rows[start:end]
        sub['schedule'] = sid
        if 'rowBreaks' in grid:
            rb = [b - start for b in grid['rowBreaks'] if start < b < end]
            if rb:
                sub['rowBreaks'] = rb
            else:
                sub.pop('rowBreaks', None)
        pages.append(sub)
    return pages


def ha_code(al):
    h = al.horizontal
    return {
        'center': 'c', 'centerContinuous': 'cc', 'right': 'r',
        'left': 'l', 'general': None,
    }.get(h, None) if h else None


def has_border(side):
    return bool(side and side.style)


def cell_borders(cell):
    b = cell.border
    return (has_border(b.top), has_border(b.bottom), has_border(b.left), has_border(b.right))


def display_text(cell):
    v = cell.value
    if v is None:
        return None
    # Year-range header cells come through as dates; the form shows only the
    # year. Everything else should mimic Excel's displayed text where the raw
    # value alone is misleading.
    if isinstance(v, (datetime.datetime, datetime.date)):
        return str(v.year)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        fmt = cell.number_format or ''
        if v < 0 and re.search(r'\([^)]*0[^)]*\)', fmt):
            n = abs(v)
            if float(n).is_integer():
                return f'({int(n):,})'
            return f'({n:,})'
    return str(v)


def extract_sheet(ws):
    maxc = ws.max_column
    # Trim trailing rows: keep through the last row carrying text.
    last = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, maxc + 1):
            if ws.cell(row=r, column=c).value is not None:
                last = r
                break
    if last == 0:
        return None
    last = min(last + 1, ws.max_row)

    # Column widths (Excel units); default ~8.43. Trim trailing empty columns.
    widths = []
    for c in range(1, maxc + 1):
        letter = openpyxl.utils.get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        widths.append(round(dim.width, 2) if dim and dim.width else 8.43)

    rows = []
    row_numbers = []
    for r in range(1, last + 1):
        rd = ws.row_dimensions.get(r)
        cells = []
        for c in range(1, maxc + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            bt, bb, bl, br = cell_borders(cell)
            if v is None and not (bt or bb or bl or br):
                continue
            o = {'c': c - 1}
            if v is not None:
                o['t'] = display_text(cell)
            f = cell.font
            if f.bold:
                o['b'] = 1
            if f.italic:
                o['i'] = 1
            if f.size and abs(f.size - 7) > 0.1:
                o['sz'] = f.size
            ha = ha_code(cell.alignment)
            if ha:
                o['ha'] = ha
            if cell.alignment.wrap_text:
                o['w'] = 1
            if cell.alignment.textRotation:
                o['tr'] = cell.alignment.textRotation
            bd = (bt and 't' or '') + (bb and 'b' or '') + (bl and 'l' or '') + (br and 'r' or '')
            if bd:
                o['bd'] = bd
            cells.append(o)
        if cells or (rd and rd.height):
            row = {'cells': cells}
            if rd and rd.height:
                row['h'] = round(rd.height, 1)
            rows.append(row)
            row_numbers.append(r)

    row_breaks = []
    for brk in ws.row_breaks.brk:
        # openpyxl break IDs are the last Excel row before a manual page break.
        # Store the corresponding zero-based row-array index where the next
        # printed page begins, so renderers do not need original Excel row IDs.
        idx = next((i for i, row_num in enumerate(row_numbers) if row_num > brk.id), None)
        if idx is not None and 0 < idx < len(rows):
            row_breaks.append(idx)

    grid = {'cols': widths, 'rows': rows}
    if row_breaks:
        grid['rowBreaks'] = sorted(set(row_breaks))
    return grid


def extract_form(path):
    """Ordered list of grid dicts (cols/rows/[rowBreaks]/sheet/schedule) for one
    form, in workbook sheet order. A combined sheet expands to one grid per
    schedule (split at each banner)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    grids = []
    for name in wb.sheetnames:
        grid = extract_sheet(wb[name])
        if grid is None:
            continue
        grid['sheet'] = name
        ids = sheet_schedules(name)
        if len(ids) > 1:
            grids.extend(_split_combined(grid, ids))
        else:
            grid['schedule'] = ids[0] if ids else None
            grids.append(grid)
    return grids


def _grid_body(grid):
    body = {'cols': grid['cols'], 'rows': grid['rows']}
    if 'rowBreaks' in grid:
        body['rowBreaks'] = grid['rowBreaks']
    return body


def _sched_key(sid):
    """Numeric sort key for a schedule id (e.g. '210A' -> (210, 'A')); non-numeric
    ids sort last."""
    m = re.match(r'(\d+)([A-Za-z]?)', sid or '')
    return (int(m.group(1)), m.group(2)) if m else (10 ** 9, sid or '')


def main():
    versions = [v for v, _ in FORMS]
    latest = versions[-1]                 # the current revision drives the page list / nav
    per_version = {}                      # version -> {schedule_id: grid}
    base_pages = None
    for v, path in FORMS:
        ordered = extract_form(path)
        per_version[v] = {g['schedule']: g for g in ordered if g['schedule']}
        if v == latest:
            base_pages = ordered           # full page set (incl. instruction pages)

    # Tag every page with the form versions whose workbook actually carries that
    # schedule, so the viewer shows a page only for filings on a matching revision
    # (front-matter pages carry no schedule and appear on every version).
    sched_versions = {}
    for v in versions:
        for sched in per_version[v]:
            sched_versions.setdefault(sched, []).append(v)
    for page in base_pages:
        page['form_versions'] = sched_versions.get(page.get('schedule'), list(versions))

    # The NAV/page list comes from the current revision. For schedules the older form
    # drew differently we keep its grid as a variant the viewer swaps in for that
    # revision. Schedules the current form DROPPED (retired pre-2016 schedules: 230,
    # 339, 416, 460, 721-726, ...) have no current page, so add their older-form page
    # directly - tagged with the versions that carry it - inserted in schedule order
    # so the nav stays in form order for legacy filings.
    variants = {}                          # schedule -> {older_version: grid body}
    legacy_only = {}                       # schedule -> grid (first older form that has it)
    for v in versions[:-1]:
        for sched, grid in per_version[v].items():
            base = per_version[latest].get(sched)
            if base is None:
                legacy_only.setdefault(sched, grid)
                continue
            if (grid['cols'], grid['rows']) != (base['cols'], base['rows']):
                variants.setdefault(sched, {})[v] = _grid_body(grid)

    for sched in sorted(legacy_only, key=_sched_key):
        grid = legacy_only[sched]
        grid['form_versions'] = sched_versions.get(sched, [])
        key = _sched_key(sched)
        at = next((i for i, p in enumerate(base_pages)
                   if p.get('schedule') and _sched_key(p['schedule']) > key), len(base_pages))
        base_pages.insert(at, grid)

    out = {'form_versions': versions, 'default_version': latest,
           'pages': base_pages, 'variants': variants}
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as fh:
        json.dump(out, fh, ensure_ascii=False, separators=(',', ':'))
    size = os.path.getsize(OUT)
    print(f'wrote {OUT}  ({size/1024:.0f} KB, {len(base_pages)} pages, '
          f'{len(legacy_only)} legacy-only pages added, '
          f'{len(variants)} schedules with older-form variants)')
    for sched in ('200', '210', '210A'):
        vs = list(variants.get(sched, {}))
        print(f'  {sched:5} variants={vs}')


if __name__ == '__main__':
    main()
